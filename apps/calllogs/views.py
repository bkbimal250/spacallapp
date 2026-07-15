"""
Views for the CallLogs app.

Endpoints:
    POST /calllogs/sync/              → Android device syncs call log batch.
    GET  /calllogs/                   → List call logs (filtered by role).
    GET  /calllogs/stats/             → Aggregate stats (total, missed, etc).
    GET  /calllogs/branch_summary/    → Per-branch call log summary.
    GET  /calllogs/export_excel/      → Download call logs as Excel file.
    POST /calllogs/bulk_delete/       → Delete multiple call logs (super_admin only).

Access Control:
    super_admin    → See all call logs across all branches.
    admin          → See all call logs across all branches.
    branch_manager → See only call logs for their assigned branch.

Android Sync Flow:
    1. Android device sends a batch of call log objects to /calllogs/sync/.
    2. Device is authenticated via HMAC (device_id + secret_key).
    3. Each call log is created (duplicate call_hash entries are silently skipped).
    4. For each newly created call log, a LeadManagement record is auto-created
       with status='pending'. This is the core lead generation logic.
    5. Device's last_sync timestamp is updated.
"""

from datetime import datetime, timedelta, timezone as datetime_timezone

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q, Sum, Avg
from rest_framework import viewsets, permissions, views, response, status, filters
from rest_framework.exceptions import ParseError
from rest_framework.decorators import action
from drf_spectacular.utils import extend_schema, OpenApiParameter, inline_serializer
from rest_framework import serializers
import openpyxl
import logging
import uuid
import hashlib
from openpyxl.styles import Font
from django.db.models import Count, Q, Sum, Avg
from django_filters.rest_framework import DjangoFilterBackend
from .filters import CallLogFilter
from .models import CallLog, MissedCallFollowUp
from .serializers import (
    CallLogSerializer,
    CallLogListSerializer,
    CallLogSyncItemSerializer,
    MissedCallFollowUpSerializer,
)
from core.authentication import DeviceAuthentication
from core.permissions import IsDevice
from apps.common.permissions import IsSuperAdmin
from apps.common.utils import apply_branch_filter
from apps.devices.services import DeviceService


logger = logging.getLogger(__name__)
FUTURE_CALL_TIME_TOLERANCE = timedelta(minutes=10)


def _remote_ip(request):
    forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def _request_id(request):
    value = request.headers.get("X-Request-ID") or getattr(request, "_device_request_id", None)
    if not value:
        value = str(uuid.uuid4())
        request._device_request_id = value
    return value


def _sync_log_context(request, device=None):
    return {
        "request_id": _request_id(request),
        "path": getattr(request, "path", ""),
        "remote_ip": _remote_ip(request),
        "device_id": getattr(device, "device_id", None),
    }


def _safe_payload_diagnostics(payloads):
    if not isinstance(payloads, list):
        return {"payload_type": type(payloads).__name__}

    keys = set()
    malformed_indices = []
    for index, item in enumerate(payloads[:25]):
        if isinstance(item, dict):
            keys.update(item.keys())
        else:
            malformed_indices.append(index)

    return {
        "payload_type": "list",
        "payload_count": len(payloads),
        "sample_keys": sorted(keys),
        "malformed_indices": malformed_indices,
    }


def _serializer_error_diagnostics(errors):
    invalid_fields = set()
    invalid_datetime_indices = []
    invalid_item_indices = []

    if isinstance(errors, list):
        for index, item_errors in enumerate(errors):
            if not item_errors:
                continue
            invalid_item_indices.append(index)
            if isinstance(item_errors, dict):
                invalid_fields.update(item_errors.keys())
                if "call_time" in item_errors:
                    invalid_datetime_indices.append(index)

    return {
        "invalid_fields": sorted(invalid_fields),
        "invalid_item_indices": invalid_item_indices[:50],
        "invalid_item_count": len(invalid_item_indices),
        "invalid_datetime_indices": invalid_datetime_indices[:50],
    }


def _safe_errors(errors):
    if isinstance(errors, list):
        return [_safe_errors(item) for item in errors]
    if isinstance(errors, dict):
        return {str(key): _safe_errors(value) for key, value in errors.items()}
    return str(errors)


def _device_call_time(item):
    call_time_ms = item.get("call_time_ms")
    if call_time_ms is not None:
        return datetime.fromtimestamp(int(call_time_ms) / 1000, tz=datetime_timezone.utc)

    call_time = item.get("call_time")
    if call_time and timezone.is_naive(call_time):
        return timezone.make_aware(call_time, datetime_timezone.utc)
    return call_time


def _server_safe_call_time(item, server_now):
    reported_call_time = _device_call_time(item)
    if reported_call_time and reported_call_time > server_now + FUTURE_CALL_TIME_TOLERANCE:
        return {
            "call_time": server_now,
            "device_reported_call_time": reported_call_time,
            "is_time_invalid": True,
            "invalid_time_reason": "future_call_time",
        }
    return {
        "call_time": reported_call_time,
        "device_reported_call_time": reported_call_time,
        "is_time_invalid": False,
        "invalid_time_reason": "",
    }


def _stable_call_hash(device, phone_number, call_time_ms, call_time, call_type, duration):
    timestamp_key = call_time_ms
    if timestamp_key is None and call_time:
        timestamp_key = int(call_time.timestamp() * 1000)
    source = "|".join([
        str(device.device_id or device.id),
        str(phone_number or ""),
        str(timestamp_key or ""),
        str(call_type or ""),
        str(duration or 0),
    ])
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def _canonical_call_hash(device, phone_number, call_time_ms, call_time, call_type, duration, sim_slot):
    import re
    clean_phone = re.sub(r"\D", "", phone_number or "")
    normalized_phone = clean_phone[-10:] if len(clean_phone) >= 10 else clean_phone
    timestamp_key = call_time_ms
    if timestamp_key is None and call_time:
        timestamp_key = int(call_time.timestamp() * 1000)
    source = "|".join([
        str(device.device_id or device.id),
        str(normalized_phone),
        str(call_type or ""),
        str(timestamp_key or ""),
        str(duration or 0),
        str(sim_slot or 1),
    ])
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def _client_call_hash(device, client_call_id):
    if not client_call_id:
        return None
    source = f"{device.device_id or device.id}|{client_call_id}"
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


class DeviceSyncView(views.APIView):
    """
    Android device batch sync endpoint.

    Authenticated via DeviceAuthentication (HMAC using device_id + secret_key).
    Accepts a JSON array of call log objects.
    Each unique call (by call_hash) is created exactly once.
    For each new call log, a Lead is auto-created with status='pending'.

    Expected payload (array of objects):
        [
          {
            "phone_number": "+919876543210",
            "call_type": "incoming",
            "duration": 120,
            "sim_slot": 0,
            "call_time": "2024-01-15T10:30:00Z",
            "call_hash": "abc123def456..."
          },
          ...
        ]
    """
    authentication_classes = [DeviceAuthentication]
    permission_classes = [IsDevice]

    def handle_exception(self, exc):
        if isinstance(exc, ParseError):
            context = _sync_log_context(self.request, getattr(self.request, "auth", None))
            logger.warning(
                "Call log sync rejected: malformed request payload",
                extra=context,
            )
            return response.Response(
                {
                    "error": "Malformed JSON payload.",
                    "code": "malformed_json",
                    "details": str(exc.detail),
                    "request_id": context["request_id"],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        return super().handle_exception(exc)

    @extend_schema(
        summary="Sync call logs (Android)",
        description="Batch upload call logs from the Android app. HMAC authentication required.",
        request=inline_serializer(
            name="CallLogSyncRequest",
            many=True,
            fields={
                "phone_number": serializers.CharField(help_text="External caller's number"),
                "call_type": serializers.ChoiceField(choices=["incoming", "outgoing", "missed", "rejected"]),
                "duration": serializers.IntegerField(help_text="Duration in seconds"),
                "sim_slot": serializers.IntegerField(help_text="Android slot: 0 or 1"),
                "call_time": serializers.DateTimeField(required=False),
                "call_time_ms": serializers.IntegerField(required=False, help_text="Raw Android CallLog.Calls.DATE epoch milliseconds"),
                "call_hash": serializers.CharField(help_text="Unique hash of the call record"),
            }
        ),
        responses={
            201: inline_serializer(
                name="CallLogSyncResponse",
                fields={
                    "status": serializers.CharField(),
                    "synced_count": serializers.IntegerField(),
                    "leads_created": serializers.IntegerField(),
                }
            )
        }
    )
    def post(self, request):
        device = request.auth  # The authenticated Device object
        log_context = _sync_log_context(request, device)

        payloads = request.data
        if not isinstance(payloads, list):
            logger.warning(
                "Call log sync rejected: payload is not a JSON array",
                extra={**log_context, **_safe_payload_diagnostics(payloads)},
            )
            return response.Response(
                {
                    "error": "Payload must be a JSON array of call log objects.",
                    "code": "invalid_payload",
                    "details": {"expected": "array", "received": type(payloads).__name__},
                    "request_id": log_context["request_id"],
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── Step 1: Pre-fetch contacts for all phone numbers in this batch ──
        # We match by last 10 digits to handle variations like +91, 0, etc.
        if not payloads:
            logger.info(
                "Call log sync received empty payload",
                extra={**log_context, "payload_count": 0},
            )

        malformed_indices = [
            index for index, item in enumerate(payloads[:50]) if not isinstance(item, dict)
        ]
        if malformed_indices:
            logger.warning(
                "Call log sync rejected: malformed items in payload",
                extra={**log_context, **_safe_payload_diagnostics(payloads)},
            )
            return response.Response(
                {
                    "error": "Each payload item must be a JSON object.",
                    "code": "malformed_payload_items",
                    "details": {"invalid_item_indices": malformed_indices},
                    "request_id": log_context["request_id"],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = CallLogSyncItemSerializer(data=payloads, many=True)
        if not serializer.is_valid():
            diagnostics = _serializer_error_diagnostics(serializer.errors)
            safe_errors = _safe_errors(serializer.errors)
            logger.warning(
                "Call log sync rejected: serializer validation failed",
                extra={
                    **log_context,
                    **_safe_payload_diagnostics(payloads),
                    **diagnostics,
                    "serializer_errors": safe_errors,
                },
            )
            return response.Response(
                {
                    "error": "Invalid call log payload.",
                    "code": "validation_error",
                    "details": serializer.errors,
                    "invalid_fields": diagnostics["invalid_fields"],
                    "request_id": log_context["request_id"],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        payloads = serializer.validated_data
        phone_numbers = {item.get("phone_number") for item in payloads if item.get("phone_number")}

        from apps.contacts.models import Contact
        contact_map = {}  # Maps last_10_digits → Contact object

        if phone_numbers:
            # Build a list of normalized numbers (last 10 digits cleaning non-digits)
            import re
            normalized_numbers = []
            for pn in phone_numbers:
                if not pn: continue
                clean_pn = re.sub(r'\D', '', pn)
                last_10 = clean_pn[-10:] if len(clean_pn) >= 10 else clean_pn
                normalized_numbers.append(last_10)
            
            # Efficiently fetch all matching contacts in one query using the indexed field
            contacts = Contact.objects.filter(phone_normalized__in=normalized_numbers)
            for c in contacts:
                contact_map[c.phone_normalized] = c

        # ── Step 2: Build CallLog objects for bulk insert ──
        logs_to_create = []
        server_now = timezone.now()
        invalid_time_count = 0
        prepared_items = []
        sync_results = []
        all_candidate_hashes = set()
        for item in payloads:
            # Normalize sim_slot: Android uses 0-indexed (0, 1) → we use 1-indexed (1, 2)
            raw_slot = item.get("sim_slot", 1)
            try:
                raw_slot = int(raw_slot)
                # Android slot 0 → SIM 1, slot 1 → SIM 2 (odd/even fallback for unusual values)
                normalized_slot = 1 if raw_slot % 2 == 0 else 2
            except (ValueError, TypeError):
                normalized_slot = 1

            phone_num = item.get("phone_number")
            import re
            clean_pn = re.sub(r'\D', '', phone_num or '')
            log_last_10 = clean_pn[-10:] if len(clean_pn) >= 10 else clean_pn
            time_values = _server_safe_call_time(item, server_now)
            if time_values["is_time_invalid"]:
                invalid_time_count += 1
            legacy_hash = _stable_call_hash(
                device=device,
                phone_number=phone_num,
                call_time_ms=item.get("call_time_ms"),
                call_time=time_values["device_reported_call_time"] or time_values["call_time"],
                call_type=item.get("call_type"),
                duration=item.get("duration", 0),
            )
            canonical_hash = _canonical_call_hash(
                device=device,
                phone_number=phone_num,
                call_time_ms=item.get("call_time_ms"),
                call_time=time_values["device_reported_call_time"] or time_values["call_time"],
                call_type=item.get("call_type"),
                duration=item.get("duration", 0),
                sim_slot=normalized_slot,
            )
            client_hash = _client_call_hash(device, item.get("client_call_id"))
            submitted_hash = (item.get("call_hash") or "").strip() or None
            candidate_hashes = [
                value for value in [client_hash, submitted_hash, legacy_hash, canonical_hash]
                if value
            ]
            all_candidate_hashes.update(candidate_hashes)
            prepared_items.append({
                "item": item,
                "phone_num": phone_num,
                "log_last_10": log_last_10,
                "normalized_slot": normalized_slot,
                "time_values": time_values,
                "canonical_hash": canonical_hash,
                "candidate_hashes": candidate_hashes,
            })

        existing_hashes = set(
            CallLog.objects.filter(call_hash__in=all_candidate_hashes).values_list("call_hash", flat=True)
        ) if all_candidate_hashes else set()

        for prepared in prepared_items:
            duplicate_hash = next(
                (candidate for candidate in prepared["candidate_hashes"] if candidate in existing_hashes),
                None,
            )
            final_hash = duplicate_hash or prepared["canonical_hash"]
            if duplicate_hash:
                sync_results.append({
                    "call_hash": final_hash,
                    "client_call_id": prepared["item"].get("client_call_id") or "",
                    "status": "already_synced",
                })
                continue

            logs_to_create.append(
                CallLog(
                    branch_id=device.branch_id,     # Inherited from device's assigned branch
                    device_id=device.id,            # The device that captured this call
                    contact=contact_map.get(prepared["log_last_10"]),  # Auto-link if known contact
                    phone_number=prepared["phone_num"],
                    phone_normalized=prepared["log_last_10"],
                    call_type=prepared["item"].get("call_type"),
                    duration=prepared["item"].get("duration", 0),
                    sim_slot=prepared["normalized_slot"],
                    call_time=prepared["time_values"]["call_time"],
                    device_reported_call_time=prepared["time_values"]["device_reported_call_time"],
                    is_time_invalid=prepared["time_values"]["is_time_invalid"],
                    invalid_time_reason=prepared["time_values"]["invalid_time_reason"],
                    call_hash=final_hash,
                )
            )
            sync_results.append({
                "call_hash": final_hash,
                "client_call_id": prepared["item"].get("client_call_id") or "",
                "status": "created",
            })

        # ── Step 3: Bulk insert — ignore duplicates (by call_hash unique constraint) ──
        if logs_to_create:
            CallLog.objects.bulk_create(logs_to_create, ignore_conflicts=True)

        submitted_hashes = [result["call_hash"] for result in sync_results]
        stored_hashes = set(
            CallLog.objects.filter(call_hash__in=submitted_hashes).values_list("call_hash", flat=True)
        ) if submitted_hashes else set()
        for result in sync_results:
            if result["status"] == "created" and result["call_hash"] not in stored_hashes:
                result["status"] = "already_synced"

        created_hashes = [
            result["call_hash"] for result in sync_results if result["status"] == "created"
        ]
        created_count = len(created_hashes)
        duplicate_count = sum(1 for result in sync_results if result["status"] == "already_synced")
        if created_hashes:
            from .services import FollowUpService
            inserted_logs = CallLog.objects.filter(call_hash__in=created_hashes)
            FollowUpService.process_batch(inserted_logs)
            
            # ── New: Hook up FollowUpService to process missed calls ──

        # ── Step 4: Auto-create Leads for newly created call logs ──
        # We query the database to find which hashes actually got inserted.
        # This avoids creating duplicate leads for already-existing call logs.
        from apps.leadmanagement.models import LeadManagement

        # Find the real IDs of inserted call logs (only the ones that are new)
        existing_lead_calllog_ids = set(
            LeadManagement.objects.filter(
                calllog__call_hash__in=created_hashes
            ).values_list("calllog_id", flat=True)
        )

        # Fetch newly created call logs that don't yet have a lead
        new_logs = CallLog.objects.filter(
            call_hash__in=created_hashes
        ).exclude(
            id__in=existing_lead_calllog_ids
        ).values("id", "contact_id", "branch_id")

        # Build lead records for bulk insert
        leads_to_create = [
            LeadManagement(
                calllog_id=log["id"],
                contact_id=log["contact_id"],
                branch_id=log["branch_id"],
                status="pending",  # All new call logs start as pending leads
            )
            for log in new_logs
        ]

        if leads_to_create:
            LeadManagement.objects.bulk_create(leads_to_create, ignore_conflicts=True)

        # ── Step 5: Update sync timestamp ──
        DeviceService.update_sync_time(device)

        # Reset monitoring flags
        from apps.monitoring.models import DeviceHealth
        DeviceHealth.objects.update_or_create(
            device=device,
            defaults={
                "last_sync": device.last_sync,
                "notified_2h": False,
                "notified_24h": False,
                "is_online": True
            }
        )

        logger.info(
            "Call log sync succeeded",
            extra={
                **log_context,
                "payload_count": len(payloads),
                "created_count": created_count,
                "duplicate_count": duplicate_count,
                "leads_created": len(leads_to_create),
                "invalid_time_count": invalid_time_count,
            },
        )
        return response.Response({
            "status": "success",
            "synced_count": len(sync_results),
            "created_count": created_count,
            "duplicate_skipped": duplicate_count,
            "leads_created": len(leads_to_create),
            "invalid_time_count": invalid_time_count,
            "items": sync_results,
        }, status=status.HTTP_201_CREATED)


class CallLogViewSet(viewsets.ModelViewSet):
    """
    Call Log CRUD and analytics viewset.

    Used by the web dashboard to view, filter, and analyze call logs.
    Access is filtered by role — branch_managers only see their branch.

    Key custom actions:
        stats          → Aggregate counts and durations.
        branch_summary → Per-branch breakdown of call types.
        export_excel   → Download all filtered logs as an Excel file.
        bulk_delete    → Delete multiple logs (super_admin only).
    """
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = CallLogFilter
    ordering_fields = ['call_time', 'duration', 'phone_number']
    ordering = ['-call_time']

    def get_serializer_class(self):
        """Use a lightweight serializer for list operations."""
        if self.action == 'list':
            return CallLogListSerializer
        return CallLogSerializer

    @extend_schema(
        summary="List/Filter Call Logs",
        description="Returns a paginated list of call logs with extensive filtering options.",
        parameters=[
            OpenApiParameter("search", str, description="Search by number or contact name"),
            OpenApiParameter("call_type", str, enum=["incoming", "outgoing", "missed", "rejected"]),
            OpenApiParameter("branch", str, description="Branch UUID"),
            OpenApiParameter("device", str, description="Device ID string"),
            OpenApiParameter("start_date", str, description="YYYY-MM-DD"),
            OpenApiParameter("end_date", str, description="YYYY-MM-DD"),
            OpenApiParameter("is_unique", bool, description="If true, returns only the latest call per phone number"),
        ],
    )
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def get_permissions(self):
        """
        Delete operations require super_admin.
        All other operations require simple authentication.
        """
        if self.action in ["destroy", "bulk_delete"]:
            return [permissions.IsAuthenticated(), IsSuperAdmin()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        """
        Return call logs filtered by the user's role and branch access.

        super_admin / admin → See ALL call logs.
        branch_manager      → See ONLY call logs for their assigned branch.
        """
        if getattr(self, "swagger_fake_view", False):
            return CallLog.objects.none()

        user = self.request.user
        
        # All views need these relations pre-fetched to avoid N+1 queries
        related = ["branch", "contact", "lead", "followup_status", "device"]
        queryset = CallLog.objects.select_related(*related).order_by("-call_time")

        # SPA manager: strict filter to single assigned branch
        if user.is_authenticated and hasattr(user, 'role') and user.role in ["spa_manager", "area_manager"]:
            if user.role == "area_manager":
                queryset = apply_branch_filter(queryset, "branch_id", user)
            elif user.branch:
                queryset = queryset.filter(branch=user.branch)
            else:
                # No branch assigned → return nothing (prevent data leak)
                return queryset.none()

        # super_admin and admin see all — no filter

        return queryset


    @extend_schema(
        summary="Bulk delete call logs",
        description="Delete a large batch of call logs at once. Restricted to super_admin.",
        request=inline_serializer(
            name="BulkDeleteRequest",
            fields={"ids": serializers.ListField(child=serializers.UUIDField())}
        ),
        responses={
            200: inline_serializer(
                name="BulkDeleteResponse",
                fields={
                    "status": serializers.CharField(),
                    "message": serializers.CharField()
                }
            )
        }
    )
    @action(detail=False, methods=["post"])
    def bulk_delete(self, request):
        """
        Bulk delete call logs by IDs. Restricted to super_admin.
        Expects: {"ids": ["uuid1", "uuid2", ...]}
        """
        ids = request.data.get("ids", [])
        if not ids:
            return response.Response(
                {"error": "No IDs provided."},
                status=status.HTTP_400_BAD_REQUEST
            )

        deleted_count, _ = CallLog.objects.filter(id__in=ids).delete()
        return response.Response({
            "status": "success",
            "message": f"Successfully deleted {deleted_count} call log(s)."
        }, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Get call log summary stats",
        description="Calculates totals and averages for the currently filtered set of logs.",
        parameters=[
            OpenApiParameter("search", str),
            OpenApiParameter("call_type", str),
            OpenApiParameter("branch", str),
            OpenApiParameter("start_date", str),
            OpenApiParameter("end_date", str),
        ],
        responses={
            200: inline_serializer(
                name="CallLogStatsResponse",
                fields={
                    "total": serializers.IntegerField(),
                    "incoming": serializers.IntegerField(),
                    "outgoing": serializers.IntegerField(),
                    "missed": serializers.IntegerField(),
                    "rejected": serializers.IntegerField(),
                    "unique_count": serializers.IntegerField(),
                    "total_duration": serializers.IntegerField(),
                    "avg_duration": serializers.FloatField(),
                }
            )
        }
    )
    @action(detail=False, methods=["get"])
    def stats(self, request):
        """
        Returns aggregate statistics for the filtered call log queryset.
        Respects role-based branch filtering.

        Response:
            total, incoming, outgoing, missed, rejected, total_duration, avg_duration
        """
        queryset = self.filter_queryset(self.get_queryset())
        stats = queryset.aggregate(
            total=Count("id"),
            incoming=Count("id", filter=Q(call_type="incoming")),
            outgoing=Count("id", filter=Q(call_type="outgoing")),
            missed=Count("id", filter=Q(call_type="missed")),
            rejected=Count("id", filter=Q(call_type="rejected")),
            total_duration=Sum("duration"),
            avg_duration=Avg("duration"),
            # Follow-up specific stats
            followed_up=Count("id", filter=Q(followup_status__is_followed_up=True)),
            sla_good=Count("id", filter=Q(followup_status__sla_status='GOOD')),
            sla_missed=Count("id", filter=Q(followup_status__sla_status='MISSED')),
            # Calculate unique count in the same pass using the indexed normalized field
            unique_count=Count("phone_normalized", distinct=True)
        )
        
        return response.Response(stats)

    @extend_schema(
        summary="Get per-branch summaries",
        description="Returns an overview of call counts for each branch.",
        parameters=[
            OpenApiParameter("branch_search", str, description="Search by branch name, city, area, spa code, phone, address, or group"),
            OpenApiParameter("city", str),
            OpenApiParameter("status", str, enum=["active", "inactive"]),
        ],
        responses={
            200: inline_serializer(
                name="BranchSummaryResponse",
                many=True,
                fields={
                    "branch_id": serializers.UUIDField(),
                    "branch_name": serializers.CharField(),
                    "city": serializers.CharField(),
                    "area": serializers.CharField(),
                    "total_calls": serializers.IntegerField(),
                    "total_missed": serializers.IntegerField(),
                    "total_outgoing": serializers.IntegerField(),
                    "total_incoming": serializers.IntegerField(),
                }
            )
        }
    )
    @action(detail=False, methods=["get"])
    def branch_summary(self, request):
        """
        Returns per-branch call log summary.
        Admins see all branches; branch managers see only their branch.

        Filters: Supports all filters from CallLogFilter (start_date, end_date, 
        quick_date, branch_search, city, status, etc).
        Sorting: Use 'ordering' parameter (e.g., -total_calls).
        """
        # Automatically applies role-based and query-params-based filtering
        queryset = self.filter_queryset(self.get_queryset())

        # Define allowed sort fields and map them to database/annotation names
        sort_mapping = {
            'total_calls': 'total_calls',
            'missed_calls': 'total_missed',
            'outgoing_calls': 'total_outgoing',
            'incoming_calls': 'total_incoming',
            'followed': 'total_followed',
            'missed_sla': 'total_missed_sla',
            'branch_name': 'branch__spa_name',
        }
    

        # Handle ordering parameter
        ordering_param = request.query_params.get('ordering', 'branch_name')
        is_desc = ordering_param.startswith('-')
        clean_field = ordering_param[1:] if is_desc else ordering_param
        
        # Determine the database field to sort by
        db_sort_field = sort_mapping.get(clean_field, 'branch__spa_name')
        if is_desc:
            db_sort_field = f'-{db_sort_field}'

        summary = queryset.order_by().values(
            "branch__id",
            "branch__spa_name",
            "branch__city",
            "branch__area",
        ).annotate(
            total_calls=Count("id"),
            total_missed=Count("id", filter=Q(call_type="missed")),
            total_outgoing=Count("id", filter=Q(call_type="outgoing")),
            total_incoming=Count("id", filter=Q(call_type="incoming")),
            total_followed=Count("id", filter=Q(followup_status__is_followed_up=True)),
            total_missed_sla=Count("id", filter=Q(followup_status__sla_status='MISSED')),
        ).order_by(db_sort_field)

        page = self.paginate_queryset(summary)
        result = self._format_branch_summary(page if page is not None else summary)

        if page is not None:
            return self.get_paginated_response(result)
        return response.Response(result, status=status.HTTP_200_OK)

    def _format_branch_summary(self, summary):
        """Helper to format branch summary queryset into clean response dicts."""
        return [
            {
                "branch_id": s["branch__id"],
                "branch_name": s["branch__spa_name"] or "Unknown Branch",
                "city": s["branch__city"] or "N/A",
                "area": s["branch__area"] or "N/A",
                "total_calls": s["total_calls"],
                "total_missed": s["total_missed"],
                "total_outgoing": s["total_outgoing"],
                "total_incoming": s["total_incoming"],
                "total_followed": s["total_followed"],
                "total_missed_sla": s["total_missed_sla"],
            }
            for s in summary
        ]

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        """
        Export call logs as an Excel (.xlsx) file.
        Applies same role-based and filter-based queryset restrictions.
        Uses .iterator() for memory efficiency with large datasets.
        """
        queryset = self.filter_queryset(self.get_queryset()).select_related("branch", "branch__branch_group", "device").iterator()

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Call Logs"

        # Column headers
        headers = [
            "Type", "Number", "Duration (s)", "SIM Slot",
            "Receiver Number", "Branch Group", "Branch", "Time"
        ]
        header_font = Font(bold=True)
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = header_font

        # Data rows — using iterator() for memory efficiency (no full queryset load)
        for row_num, log in enumerate(queryset, 2):
            # Determine the receiver's SIM number based on which SIM handled the call
            receiver = "N/A"
            if log.device:
                if log.sim_slot == 1:
                    receiver = log.device.sim_1_number or "N/A"
                elif log.sim_slot == 2:
                    receiver = log.device.sim_2_number or "N/A"

            worksheet.cell(row=row_num, column=1).value = log.call_type
            worksheet.cell(row=row_num, column=2).value = log.phone_number
            worksheet.cell(row=row_num, column=3).value = log.duration
            worksheet.cell(row=row_num, column=4).value = f"SIM {log.sim_slot}"
            worksheet.cell(row=row_num, column=5).value = receiver
            worksheet.cell(row=row_num, column=6).value = log.branch.branch_group.name if log.branch and log.branch.branch_group else "N/A"
            worksheet.cell(row=row_num, column=7).value = log.branch.spa_name if log.branch else "N/A"
            worksheet.cell(row=row_num, column=8).value = (
                log.call_time.strftime("%Y-%m-%d %H:%M:%S") if log.call_time else "N/A"
            )

        # Build HTTP response with Excel content type
        http_response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        http_response["Content-Disposition"] = f'attachment; filename="call_logs_{timestamp}.xlsx"'
        workbook.save(http_response)
        return http_response


class MissedCallFollowUpViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for tracking and reporting on missed call follow-up performance.
    """
    queryset = MissedCallFollowUp.objects.all()
    serializer_class = MissedCallFollowUpSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset().select_related('missed_call', 'branch', 'followup_call')
        
        return apply_branch_filter(qs, "branch_id", user)

    @action(detail=False, methods=['get'])
    def today_missed(self, request):
        """Returns all missed calls received today."""
        today = timezone.now().date()
        queryset = self.get_queryset().filter(created_at__date=today)
        serializer = self.get_serializer(queryset, many=True)
        return response.Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Returns aggregate stats: Followed vs Not Followed, and SLA breakdown."""
        qs = self.get_queryset()
        
        # Optimized aggregation for millions of records
        stats = qs.aggregate(
            total_missed=Count('id'),
            followed_count=Count('id', filter=Q(is_followed_up=True)),
            unfollowed_count=Count('id', filter=Q(is_followed_up=False)),
            good_sla=Count('id', filter=Q(sla_status='GOOD')),
            ok_sla=Count('id', filter=Q(sla_status='OK')),
            late_sla=Count('id', filter=Q(sla_status='LATE')),
            missed_sla=Count('id', filter=Q(sla_status='MISSED')),
        )
        
        return response.Response(stats)

    @action(detail=False, methods=['get'])
    def branch_performance(self, request):
        """Returns missed call follow-up performance breakdown per branch."""
        # Note: In a large-scale system, this should be cached or pre-aggregated in a DailySummary model.
        performance = self.get_queryset().values('branch__spa_name').annotate(
            total=Count('id'),
            followed=Count('id', filter=Q(is_followed_up=True)),
            follow_rate=Count('id', filter=Q(is_followed_up=True)) * 100.0 / Count('id'),
            avg_attempts=Avg('followup_attempt_count')
        ).order_by('-follow_rate')
        
        return response.Response(performance)
